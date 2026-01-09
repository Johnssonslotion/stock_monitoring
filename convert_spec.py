#!/usr/bin/env python3
"""
KIS Excel spec을 Markdown으로 변환
"""
import sys
import openpyxl

def convert_excel_to_markdown(excel_path, output_path):
    wb = openpyxl.load_workbook(excel_path)
    
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write("# KIS 국내주식 실시간호가 스펙 (H0STCNT0)\n\n")
        f.write("> 원본: 국내주식 실시간호가 (KRX) [실시간-004].xlsx\n\n")
        
        for sheet_name in wb.sheetnames:
            f.write(f"## {sheet_name}\n\n")
            
            sheet = wb[sheet_name]
            
            # 헤더 찾기
            headers = []
            for cell in sheet[1]:
                if cell.value:
                    headers.append(str(cell.value))
            
            if headers:
                # 마크다운 테이블
                f.write("| " + " | ".join(headers) + " |\n")
                f.write("| " + " | ".join(["---"] * len(headers)) + " |\n")
                
                # 데이터 행
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if any(row):  # 빈 행 건너뛰기
                        row_data = [str(cell) if cell is not None else "" for cell in row[:len(headers)]]
                        f.write("| " + " | ".join(row_data) + " |\n")
            else:
                # 테이블 형식이 아닌 경우
                for row in sheet.iter_rows(values_only=True):
                    if any(row):
                        f.write(" | ".join([str(cell) if cell is not None else "" for cell in row]) + "\n")
            
            f.write("\n---\n\n")
    
    print(f"✅ Converted: {output_path}")

if __name__ == "__main__":
    excel_path = "docs/국내주식 실시간호가 (KRX) [실시간-004].xlsx"
    output_path = "docs/KIS_H0STCNT0_spec.md"
    convert_excel_to_markdown(excel_path, output_path)
