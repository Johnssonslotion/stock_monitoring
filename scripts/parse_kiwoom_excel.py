#!/usr/bin/env python3
import openpyxl
import json
import sys
import os

# Check file exists
filepath = 'docs/키움 REST API 문서.xlsx'
if not os.path.exists(filepath):
    print(f"❌ File not found: {filepath}")
    print(f"Current dir: {os.getcwd()}")
    print(f"Files in docs/:")
    for f in os.listdir('docs/'):
        if 'xlsx' in f.lower() or 'kiwoom' in f.lower():
            print(f"  - {f}")
    sys.exit(1)

wb = openpyxl.load_workbook(filepath)
print("📊 Sheets:", wb.sheetnames)
print()

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"Sheet: {sheet_name}")
    print(f"Dimensions: {sheet.dimensions}")
    print(f"{'='*80}\n")
    
    # Print all rows
    for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
        if i > 150:
            print(f"... (total {sheet.max_row} rows)")
            break
        if any(cell is not None for cell in row):
            print(f"{i:3d}: {row}")
    
    print()
