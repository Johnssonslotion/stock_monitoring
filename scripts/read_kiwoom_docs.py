#!/usr/bin/env python3
import openpyxl
import sys

try:
    wb = openpyxl.load_workbook('/home/ubuntu/workspace/stock_backtest/docs/키움 REST API 문서.xlsx')
    print("📊 Available Sheets:", wb.sheetnames)
    print()
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\n{'='*80}")
        print(f"📄 Sheet: {sheet_name}")
        print(f"   Dimensions: {sheet.dimensions}")
        print(f"{'='*80}\n")
        
        # Print all rows (limit to 50 per sheet)
        for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if i > 50:
                print(f"... (truncated, total rows: {sheet.max_row})")
                break
            # Filter out empty rows
            if any(cell is not None for cell in row):
                print(f"Row {i:3d}: {row}")
        
        print("\n")
            
except Exception as e:
    print(f"❌ Error: {e}", file=sys.stderr)
    import traceback
    traceback.print_exc()
